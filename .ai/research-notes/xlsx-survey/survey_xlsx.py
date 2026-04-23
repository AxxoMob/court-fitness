import openpyxl
from openpyxl import load_workbook

FILES = [
    r"C:\xampp\htdocs\ltat-fitness-module\helpful-project-documents\Big Picture Programming - 5 Day (Full - KG's).xlsx",
    r"C:\xampp\htdocs\ltat-fitness-module\helpful-project-documents\Big Picture Programming - 5 Day (Full - LBs).xlsx",
    r"C:\xampp\htdocs\ltat-fitness-module\helpful-project-documents\training load data collection.xlsm",
]

for path in FILES:
    print("=" * 80)
    print("FILE:", path.split("\\")[-1])
    print("=" * 80)
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print("  LOAD ERROR:", e)
        continue
    print("SHEETS:", wb.sheetnames)
    for sname in wb.sheetnames:
        try:
            ws = wb[sname]
            print(f"\n--- Sheet: {sname!r}  (dim={ws.dimensions})  max_row={ws.max_row} max_col={ws.max_column} ---")
        except Exception as e:
            print(f"\n--- Sheet: {sname!r}  ERROR: {e}")
    wb.close()
    print()
