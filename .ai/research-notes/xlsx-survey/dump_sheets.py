from openpyxl import load_workbook

def dump(path, sheets_to_dump, max_cols=20, max_rows=80):
    print("\n" + "=" * 90)
    print("FILE:", path.split("\\")[-1])
    print("=" * 90)
    wb = load_workbook(path, data_only=True)
    for sname in sheets_to_dump:
        if sname not in wb.sheetnames:
            print(f"\n  [{sname}] NOT FOUND")
            continue
        ws = wb[sname]
        print(f"\n--- Sheet: {sname!r}  max_row={ws.max_row}  max_col={ws.max_column} ---")
        r = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows),
                                min_col=1, max_col=min(ws.max_column, max_cols),
                                values_only=True):
            r += 1
            if all(v is None for v in row):
                continue
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                else:
                    s = str(v).replace("\n", " / ")
                    if len(s) > 60: s = s[:60] + "…"
                    cells.append(s)
            print(f"  R{r:>3}:", " | ".join(cells))
    wb.close()

kg = r"C:\xampp\htdocs\ltat-fitness-module\helpful-project-documents\Big Picture Programming - 5 Day (Full - KG's).xlsx"
tl = r"C:\xampp\htdocs\ltat-fitness-module\helpful-project-documents\training load data collection.xlsm"

dump(kg, ["Start Here  Guide", "Assessment", "The Big Picture"], max_cols=25, max_rows=100)
dump(tl, ["How to Guide", "Control Panel", "Data Entry & Manipulation", "Training Load Data"], max_cols=25, max_rows=60)
